"""Regression checks for exact banner and topline export behavior."""

from __future__ import annotations

from io import BytesIO
from unittest.mock import patch
import unittest

import pandas as pd
from openpyxl import load_workbook

from app.pages.page_12_export import (
    _build_export_filename,
    _download_signature_for_generated_workbook,
    _is_local_codex_testing,
    _local_testing_download_path,
    _is_generated_workbook_stale,
    _project_settings_download_key,
)
from src.exporter import export_workbook_to_excel_bytes
from src.tables import (
    PERFORMANCE_SIG_TEST_NOTE,
    TOPLINE_NOTES_UNAVAILABLE_NOTE,
    generate_workbook_package,
)


CREATORS = [f"Creator {index}" for index in range(1, 11)]
TIERS = ["Mega", "Macro", "Mid-Tier", "Micro"]
SCALE_VALUES = ["Love it", "Like it", "Neutral", "Dislike", "Hate it"]


def _base_dataframe(row_count: int = 400) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "content_variant": CREATORS[index % len(CREATORS)],
                "tier": TIERS[index % len(TIERS)],
                "Post_Affinity": SCALE_VALUES[index % len(SCALE_VALUES)],
            }
            for index in range(row_count)
        ]
    )


def _question_metadata() -> list[dict[str, object]]:
    return [
        {
            "variable": "content_variant",
            "display_variable_name": "Creator",
            "question_label": "Creator",
            "detected_type": "Single-Select",
            "answer_choices_list": CREATORS,
            "include": True,
        },
        {
            "variable": "tier",
            "display_variable_name": "Tier",
            "question_label": "Tier",
            "detected_type": "Single-Select",
            "answer_choices_list": TIERS,
            "include": True,
        },
        {
            "variable": "Post_Affinity",
            "display_variable_name": "Affinity",
            "question_label": "Affinity",
            "detected_type": "Scale / Likert",
            "answer_choices_list": SCALE_VALUES,
            "include": True,
        },
    ]


def _scale_mappings() -> dict[str, dict[str, object]]:
    return {
        "Post_Affinity": {
            "rows": [
                {"response_value": value, "bucket": index + 1}
                for index, value in enumerate(SCALE_VALUES)
            ]
        }
    }


def _comparison_scheme() -> dict[str, object]:
    return {
        "enabled": True,
        "mode": "exclusive",
        "control_group_id": "",
        "groups": [
            {
                "id": f"group_{index}",
                "label": creator,
                "role": "test",
                "match_logic": "ALL",
                "conditions": [
                    {
                        "variable": "content_variant",
                        "operator": "Is exactly",
                        "values": [creator],
                    }
                ],
            }
            for index, creator in enumerate(CREATORS, start=1)
        ],
    }


def _stat_config(include_lift: bool = False) -> dict[str, object]:
    return {
        "enabled": True,
        "comparison_scope": "lowest_banner_level",
        "confidence_intervals": [95],
        "include_percentage": True,
        "include_n_count": False,
        "include_lift": include_lift,
        "notation_location": "below_metric",
    }


def _topline_config(include_lift: bool = False) -> dict[str, object]:
    return {
        "configured": True,
        "variables": ["Post_Affinity"],
        "response_selections": {"Post_Affinity": ["T2B"]},
        "note_base_sections": {"Post_Affinity": "Total Answering"},
        "include_lift": include_lift,
        "comparison_scope": "lowest_banner_level",
        "include_significance_notes": True,
    }


def _generate_package(
    banner_config: dict[str, object],
    question_metadata: list[dict[str, object]] | None = None,
    banner_stat_config: dict[str, object] | None = None,
    topline_config: dict[str, object] | None = None,
) -> dict[str, object]:
    return generate_workbook_package(
        cleaned_df=_base_dataframe(),
        question_metadata=question_metadata or _question_metadata(),
        custom_variables=[],
        banner_config=banner_config,
        adhoc_crosstabs_config={"tables": []},
        net_definitions={"Post_Affinity": {"T2B": True, "T3B": False, "B2B": False, "B3B": False}},
        scale_mappings=_scale_mappings(),
        banner_stat_config=banner_stat_config or _stat_config(),
        adhoc_stat_config=_stat_config(),
        comparison_col="content_variant",
        comparison_group_order={},
        comparison_group_labels={},
        comparison_scheme=_comparison_scheme(),
        global_filters={"rows": []},
        weighting_config={"weights": []},
        topline_config=topline_config or _topline_config(),
    )


class ExportBehaviorTest(unittest.TestCase):
    def test_export_filename_strips_browser_unsafe_characters(self) -> None:
        filename = _build_export_filename("Brand / Test: Wave 1.xlsx")

        self.assertNotIn("/", filename)
        self.assertNotIn(":", filename)
        self.assertTrue(filename.endswith(".xlsx"))

    def test_stale_workbook_remains_downloadable_against_generated_signature(self) -> None:
        self.assertTrue(_is_generated_workbook_stale("generated", "current"))
        self.assertEqual(
            _download_signature_for_generated_workbook("generated", "current"),
            "generated",
        )

    def test_local_testing_downloads_are_codex_gated(self) -> None:
        with patch.dict("os.environ", {}, clear=True):
            self.assertFalse(_is_local_codex_testing())
        with patch.dict("os.environ", {"CODEX_THREAD_ID": "thread"}, clear=True):
            self.assertTrue(_is_local_codex_testing())
        with patch.dict("os.environ", {"CODEX_THREAD_ID": "thread", "BLS_LOCAL_TEST_DOWNLOADS": "0"}, clear=True):
            self.assertFalse(_is_local_codex_testing())
        with patch.dict("os.environ", {"BLS_LOCAL_TEST_DOWNLOADS": "1"}, clear=True):
            self.assertTrue(_is_local_codex_testing())

    def test_local_testing_download_path_stays_in_testing_folder(self) -> None:
        path = _local_testing_download_path("../Brand / Test: Wave 1.xlsx")

        self.assertIn("exports/local_testing", str(path))
        self.assertEqual(path.name, "Brand _ Test_ Wave 1.xlsx")

    def test_project_settings_download_key_ignores_saved_at(self) -> None:
        first_snapshot = (
            '{"kind":"bls_smart_tables_project_settings","saved_at":"2026-06-12T10:00:00Z",'
            '"project_config":{"variables":{"included_columns":["QID1"]}}}'
        )
        second_snapshot = (
            '{"kind":"bls_smart_tables_project_settings","saved_at":"2026-06-12T10:00:01Z",'
            '"project_config":{"variables":{"included_columns":["QID1"]}}}'
        )
        changed_snapshot = (
            '{"kind":"bls_smart_tables_project_settings","saved_at":"2026-06-12T10:00:01Z",'
            '"project_config":{"variables":{"included_columns":["QID2"]}}}'
        )

        first_key = _project_settings_download_key("local_testing_project_settings_download", first_snapshot)
        second_key = _project_settings_download_key("local_testing_project_settings_download", second_snapshot)
        changed_key = _project_settings_download_key("local_testing_project_settings_download", changed_snapshot)

        self.assertEqual(first_key, second_key)
        self.assertNotEqual(first_key, changed_key)

    def test_exact_single_level_banners_and_topline_net_filter(self) -> None:
        package = _generate_package(
            {
                "banners": [
                    {"name": "Creators", "level_1": "__comparison_scheme__", "level_2": "", "level_3": ""},
                    {"name": "Tier", "level_1": "tier", "level_2": "", "level_3": ""},
                ],
                "include_total": True,
                "export_style": "one_per_sheet",
            }
        )

        self.assertEqual([sheet.name for sheet in package["sheets"]], ["Creators", "Tier"])
        creators_sheet, tier_sheet = package["sheets"]
        self.assertEqual(creators_sheet.levels, ["content_variant"])
        self.assertEqual(tier_sheet.levels, ["tier"])
        self.assertEqual(len(creators_sheet.groups), 11)
        self.assertEqual(len(tier_sheet.groups), 5)
        self.assertEqual({row["Variable"] for row in package["topline_sheet"].rows}, {"Post_Affinity"})
        self.assertEqual({row["Response"] for row in package["topline_sheet"].rows}, {"T2B"})
        self.assertIn(TOPLINE_NOTES_UNAVAILABLE_NOTE, package["topline_sheet"].footnotes)

        workbook = load_workbook(BytesIO(export_workbook_to_excel_bytes(package)), read_only=True)
        self.assertEqual(workbook.sheetnames, ["Topline", "Creators", "Tier"])
        all_values = [
            value
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        ]
        self.assertFalse(any("Lift" in str(value) for value in all_values))
        topline_header_values = [
            workbook["Topline"].cell(row=11, column=column).value
            for column in range(1, workbook["Topline"].max_column + 1)
        ]
        self.assertIn("Creator 1", topline_header_values)
        self.assertIn("Creator 10", topline_header_values)

    def test_banner_tables_follow_question_metadata_order(self) -> None:
        metadata = [
            _question_metadata()[2],
            _question_metadata()[1],
            _question_metadata()[0],
        ]

        package = _generate_package(
            {
                "banners": [
                    {"name": "Creators", "level_1": "__comparison_scheme__", "level_2": "", "level_3": ""},
                ],
                "include_total": True,
                "export_style": "one_per_sheet",
            },
            question_metadata=metadata,
        )

        self.assertEqual(
            [table.variable for table in package["sheets"][0].tables],
            ["Post_Affinity", "tier", "content_variant"],
        )

    def test_explicit_nested_banner_scopes_all_cell_pairs(self) -> None:
        package = _generate_package(
            {
                "banners": [
                    {"name": "Tier x Creator", "level_1": "tier", "level_2": "__comparison_scheme__", "level_3": ""},
                ],
                "include_total": True,
                "export_style": "one_per_sheet",
            }
        )
        sheet = package["sheets"][0]
        self.assertEqual(sheet.levels, ["tier", "content_variant"])
        self.assertEqual(len(sheet.groups), 41)
        self.assertEqual(len(sheet.comparison_pairs), 4 * 45)
        self.assertEqual(package["export_summary"]["topline_notes_warning"], "")

    def test_large_sig_workload_marks_only_affected_sheet(self) -> None:
        choices = [f"Choice {index}" for index in range(200)]
        df = _base_dataframe()
        df["big_q"] = [choices[index % len(choices)] for index in range(len(df))]
        metadata = [
            {
                "variable": "big_q",
                "question_label": "Big Q",
                "detected_type": "Single-Select",
                "answer_choices_list": choices,
                "include": True,
            }
        ]
        package = generate_workbook_package(
            cleaned_df=df,
            question_metadata=metadata,
            custom_variables=[],
            banner_config={
                "banners": [
                    {"name": "Tier x Creator", "level_1": "tier", "level_2": "__comparison_scheme__", "level_3": ""},
                ],
                "include_total": True,
                "export_style": "one_per_sheet",
            },
            adhoc_crosstabs_config={"tables": []},
            net_definitions={},
            scale_mappings={},
            banner_stat_config=_stat_config(),
            adhoc_stat_config=_stat_config(),
            comparison_col="content_variant",
            comparison_group_order={},
            comparison_group_labels={},
            comparison_scheme=_comparison_scheme(),
            topline_config={
                "configured": True,
                "variables": ["big_q"],
                "response_selections": {"big_q": choices[:1]},
                "include_significance_notes": True,
            },
        )

        sheet = package["sheets"][0]
        self.assertGreaterEqual(sheet.estimated_sig_tests, 50_000)
        self.assertTrue(sheet.optimized_significance)
        self.assertIn(PERFORMANCE_SIG_TEST_NOTE, sheet.footnotes)
        self.assertEqual(package["export_summary"]["optimized_sig_sheets"], ["Tier x Creator"])

    def test_sig_header_letters_start_at_a_when_total_is_hidden(self) -> None:
        package = _generate_package(
            {
                "banners": [
                    {"name": "Creators", "level_1": "__comparison_scheme__", "level_2": "", "level_3": ""},
                ],
                "include_total": False,
                "export_style": "one_per_sheet",
            }
        )

        workbook = load_workbook(BytesIO(export_workbook_to_excel_bytes(package)), read_only=True)
        worksheet = workbook["Creators"]
        left_sig_letters = [
            worksheet.cell(row=8, column=column).value
            for column in range(3, 13)
        ]
        right_sig_letters = [
            worksheet.cell(row=8, column=column).value
            for column in range(14, 24)
        ]

        self.assertEqual(left_sig_letters, list("ABCDEFGHIJ"))
        self.assertEqual(right_sig_letters, list("ABCDEFGHIJ"))
        self.assertNotIn("@", left_sig_letters + right_sig_letters)

    def test_topline_lift_columns_do_not_require_banner_lift(self) -> None:
        package = _generate_package(
            {
                "banners": [
                    {"name": "Creators", "level_1": "__comparison_scheme__", "level_2": "", "level_3": ""},
                ],
                "include_total": True,
                "export_style": "one_per_sheet",
            },
            banner_stat_config=_stat_config(include_lift=False),
            topline_config=_topline_config(include_lift=True),
        )

        self.assertFalse(package["include_lift"])
        workbook = load_workbook(BytesIO(export_workbook_to_excel_bytes(package)), read_only=True)
        topline_header_values = [
            workbook["Topline"].cell(row=11, column=column).value
            for column in range(1, workbook["Topline"].max_column + 1)
        ]

        self.assertIn("Creator 2 vs Creator 1 Lift", topline_header_values)

    def test_banner_lift_setting_adds_lift_columns(self) -> None:
        package = _generate_package(
            {
                "banners": [
                    {"name": "Creators", "level_1": "__comparison_scheme__", "level_2": "", "level_3": ""},
                ],
                "include_total": True,
                "export_style": "one_per_sheet",
            },
            banner_stat_config=_stat_config(include_lift=True),
        )

        self.assertTrue(package["include_lift"])
        workbook = load_workbook(BytesIO(export_workbook_to_excel_bytes(package)), read_only=True)
        creator_header_values = [
            workbook["Creators"].cell(row=6, column=column).value
            for column in range(1, workbook["Creators"].max_column + 1)
        ]

        self.assertTrue(any("Lift" in str(value) for value in creator_header_values if value))


if __name__ == "__main__":
    unittest.main()
