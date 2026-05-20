"""Excel export helpers for branded BLS Smart Tables workbooks."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from src.comparisons import COMPARISON_SCHEME_DISPLAY_NAME
from src.utils import normalize_text


VN_BLACK = "FF000000"
VN_WHITE = "FFFFFFFF"
VN_RED = "FFFF005C"
VN_ORANGE = "FFFF6927"
VN_YELLOW = "FFFFC227"
VN_LIGHT_GRAY = "FFF4F5F8"
VN_BORDER_GRAY = "FFD9D9D9"
VN_GREEN = "FF1F8F4E"
VN_LIGHT_GREEN = "FFE6F4EA"
VN_LIGHT_RED = "FFFDE8EC"
EXPORT_LAYOUT_VERSION = "Layout v2026.05.19.1"


def _excel_rgb(color: str) -> str:
    """Return an 8-digit ARGB color string for openpyxl/Excel styles."""
    normalized = (color or "").strip().lstrip("#").upper()
    if len(normalized) == 6:
        return f"FF{normalized}"
    return normalized


def _apply_header_style(cell, fill_color: str, font_color: str = VN_WHITE, bold: bool = True) -> None:
    """Apply a consistent header style to one Excel cell."""
    cell.fill = PatternFill("solid", fgColor=_excel_rgb(fill_color))
    cell.font = Font(color=_excel_rgb(font_color), bold=bold, name="Proxima Nova")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_body_style(cell, bold: bool = False, fill_color: str | None = None, wrap: bool = False) -> None:
    """Apply a consistent body style to one Excel cell."""
    if fill_color:
        cell.fill = PatternFill("solid", fgColor=_excel_rgb(fill_color))
    cell.font = Font(color=_excel_rgb(VN_BLACK), bold=bold, name="Proxima Nova")
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    thin = Side(style="thin", color=_excel_rgb(VN_BORDER_GRAY))
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply_topline_delta_style(lift_cell, significant_direction: str) -> None:
    """Highlight topline test results to make wins/losses easy to scan."""
    if significant_direction == "right":
        lift_cell.fill = PatternFill("solid", fgColor=_excel_rgb(VN_LIGHT_GREEN))
        lift_cell.font = Font(color=_excel_rgb(VN_GREEN), bold=True, name="Proxima Nova")
    elif significant_direction == "left":
        lift_cell.fill = PatternFill("solid", fgColor=_excel_rgb(VN_LIGHT_RED))
        lift_cell.font = Font(color=_excel_rgb(VN_RED), bold=True, name="Proxima Nova")


def _write_version_stamp(worksheet, row: int, start_column: int, end_column: int) -> None:
    """Write a visible exporter layout version stamp."""
    if end_column <= start_column:
        stamp_cell = worksheet.cell(row=row, column=start_column, value=EXPORT_LAYOUT_VERSION)
    else:
        worksheet.merge_cells(
            start_row=row,
            start_column=start_column,
            end_row=row,
            end_column=end_column,
        )
        stamp_cell = worksheet.cell(row=row, column=start_column, value=EXPORT_LAYOUT_VERSION)
    _apply_body_style(stamp_cell, bold=True, fill_color=VN_LIGHT_GRAY)
    stamp_cell.alignment = Alignment(horizontal="right", vertical="center")


def _set_sheet_columns(worksheet, group_count: int, lift_count: int = 0) -> None:
    """Set practical column widths for a question-table worksheet."""
    worksheet.column_dimensions["A"].width = 42
    worksheet.column_dimensions["B"].width = 20
    total_data_columns = max(group_count, 1) * 2
    total_sheet_columns = total_data_columns + (lift_count * 2) + 1
    for column_index in range(3, 3 + total_sheet_columns):
        worksheet.column_dimensions[get_column_letter(column_index)].width = 14
    separator_column = 3 + group_count + lift_count
    worksheet.column_dimensions[get_column_letter(separator_column)].width = 3


def _build_banner_lift_pairs(sheet, visible_groups: list[dict]) -> tuple[list[dict], str]:
    """Return binary lowest-level lift pairs for one banner sheet."""
    configured_pairs = list(getattr(sheet, "comparison_pairs", []) or [])
    if configured_pairs:
        lift_pairs: list[dict] = []
        for pair in configured_pairs:
            subgroup_label = normalize_text(pair.get("subgroup_label"))
            left_label = normalize_text(pair.get("left_label")) or "Control"
            right_label = normalize_text(pair.get("right_label")) or "Test"
            parent_label = f"{right_label} vs {left_label}"
            if subgroup_label and subgroup_label != "Total":
                parent_label = f"{subgroup_label}: {parent_label}"
            lift_pairs.append(
                {
                    "parent_label": parent_label,
                    "left_index": int(pair.get("left_index", 0)),
                    "right_index": int(pair.get("right_index", 1)),
                }
            )
        return lift_pairs, ""

    non_total_groups = [group for group in visible_groups if group.get("label") != "Total"]
    has_explicit_comparison_roles = any(normalize_text(group.get("role")) for group in non_total_groups)
    has_control_group = any(
        normalize_text(group.get("role")).lower() == "control"
        or normalize_text(group.get("label")).casefold() == "control"
        for group in non_total_groups
    )
    if has_explicit_comparison_roles and not has_control_group:
        return [], ""

    if not sheet.levels:
        return [], "Lift could not be performed since the comparison variable is not binary."

    lowest_level = sheet.levels[-1]
    if not non_total_groups:
        return [], "Lift could not be performed since the comparison variable is not binary."

    parent_lookup: dict[tuple[str, ...], list[tuple[int, dict]]] = {}
    for group_index, group in enumerate(visible_groups):
        if group.get("label") == "Total":
            continue
        display_values = group.get("display_values", {}) or {}
        parent_key = tuple(
            normalize_text(display_values.get(level) or group.get("values", {}).get(level))
            for level in sheet.levels[:-1]
        )
        parent_lookup.setdefault(parent_key, []).append((group_index, group))

    lift_pairs: list[dict] = []
    for parent_key, members in parent_lookup.items():
        if len(members) != 2:
            return [], "Lift could not be performed since the comparison variable is not binary."
        (left_index, left_group), (right_index, right_group) = members
        left_display = left_group.get("display_values", {}) or {}
        right_display = right_group.get("display_values", {}) or {}
        lowest_left = normalize_text(left_display.get(lowest_level) or left_group.get("values", {}).get(lowest_level))
        lowest_right = normalize_text(right_display.get(lowest_level) or right_group.get("values", {}).get(lowest_level))
        if not lowest_left or not lowest_right or lowest_left == lowest_right:
            return [], "Lift could not be performed since the comparison variable is not binary."
        parent_label = " | ".join(value for value in parent_key if value)
        if not parent_label:
            parent_label = f"{lowest_right} vs {lowest_left}"
        lift_pairs.append(
            {
                "parent_label": parent_label,
                "left_index": left_index,
                "right_index": right_index,
            }
        )
    return lift_pairs, ""


def _format_lift_display(left_pct: float | None, right_pct: float | None) -> str:
    """Format a lift value as signed point difference."""
    if left_pct is None or right_pct is None:
        return ""
    lift_points = round((right_pct - left_pct) * 100)
    return f"{lift_points:+d} pts"


def _format_level_label(level_name: str, level_labels: dict[str, str] | None = None) -> str:
    """Return a human-friendly banner-level label for export headers."""
    normalized = normalize_text(level_name)
    if normalized == "__selection_status__":
        return "Selection Status"
    if normalized == "__comparison_scheme__":
        return COMPARISON_SCHEME_DISPLAY_NAME
    level_labels = level_labels or {}
    return level_labels.get(level_name, level_name)


def _paint_separator_column(worksheet, column_index: int, start_row: int, end_row: int) -> None:
    """Fill a narrow black separator column between table sections."""
    if end_row < start_row:
        return
    for row_index in range(start_row, end_row + 1):
        cell = worksheet.cell(row=row_index, column=column_index)
        if isinstance(cell, MergedCell):
            continue
        cell.fill = PatternFill("solid", fgColor=_excel_rgb(VN_BLACK))
        cell.border = Border()


def _set_topline_columns(worksheet, group_count: int = 2, pair_count: int = 0) -> None:
    """Set practical column widths for the topline worksheet."""
    worksheet.column_dimensions["A"].width = 4
    worksheet.column_dimensions["B"].width = 42
    worksheet.column_dimensions["C"].width = 15
    current_column = 4
    if pair_count:
        for _ in range(pair_count):
            worksheet.column_dimensions[get_column_letter(current_column)].width = 15
            worksheet.column_dimensions[get_column_letter(current_column + 1)].width = 18
            current_column += 2
    else:
        for _ in range(max(group_count - 1, 0)):
            worksheet.column_dimensions[get_column_letter(current_column)].width = 15
            current_column += 1
    worksheet.column_dimensions[get_column_letter(current_column)].width = 18
    note_column_count = max(pair_count, 1)
    for column_index in range(current_column + 1, current_column + 1 + note_column_count):
        worksheet.column_dimensions[get_column_letter(column_index)].width = 48
    worksheet.column_dimensions[get_column_letter(current_column + note_column_count + 1)].width = 8


def _coerce_topline_index(value, fallback: int) -> int:
    """Return a stable numeric topline group index."""
    try:
        return int(value)
    except (TypeError, ValueError):
        return fallback


def _normalize_topline_groups(group_results: list[dict]) -> list[dict]:
    """Ensure topline group payloads have positional indexes."""
    normalized_groups: list[dict] = []
    for position, group in enumerate(group_results):
        normalized_group = dict(group)
        normalized_group["index"] = _coerce_topline_index(normalized_group.get("index"), position)
        normalized_groups.append(normalized_group)
    return normalized_groups


def _topline_note_key(left_label: str, right_label: str) -> str:
    """Return the pair-key used by the topline table service."""
    return f"{normalize_text(left_label).casefold()}||{normalize_text(right_label).casefold()}"


def _normalize_topline_pairs(pair_results: list[dict], group_results: list[dict]) -> list[dict]:
    """Ensure topline comparison pairs have indexes, labels, and note keys."""
    normalized_pairs: list[dict] = []
    group_lookup = {group.get("index"): group for group in group_results}
    group_indexes = [group.get("index") for group in group_results]
    default_left_index = group_indexes[0] if group_indexes else 0
    for position, pair in enumerate(pair_results):
        fallback_right_index = (
            group_indexes[position + 1]
            if position + 1 < len(group_indexes)
            else group_indexes[1]
            if len(group_indexes) > 1
            else 1
        )
        left_index = _coerce_topline_index(pair.get("left_index"), default_left_index)
        right_index = _coerce_topline_index(pair.get("right_index"), fallback_right_index)
        left_label = normalize_text(pair.get("left_label")) or normalize_text(
            group_lookup.get(left_index, {}).get("label")
        ) or "Control"
        right_label = normalize_text(pair.get("right_label")) or normalize_text(
            group_lookup.get(right_index, {}).get("label")
        ) or f"Group {position + 2}"
        normalized_pair = dict(pair)
        normalized_pair.update(
            {
                "left_index": left_index,
                "right_index": right_index,
                "left_label": left_label,
                "right_label": right_label,
                "note_key": pair.get("note_key") or _topline_note_key(left_label, right_label),
            }
        )
        normalized_pairs.append(normalized_pair)
    return normalized_pairs


def _write_topline_sheet(workbook, topline_sheet) -> None:
    """Write the flat topline worksheet.

    Inputs:
        workbook: OpenPyXL workbook object.
        topline_sheet: Flat topline payload built by the table service.

    Outputs:
        Adds one `Topline` worksheet to the workbook.
    """
    worksheet = workbook.create_sheet(title="Topline")
    rows = list(getattr(topline_sheet, "rows", []))
    first_row = rows[0] if rows else {}
    group_results = _normalize_topline_groups(list(first_row.get("Group Results", []) or []))
    if not group_results and rows:
        group_results = _normalize_topline_groups(
            [
                {
                    "label": first_row.get("Left Label", "Group 1"),
                    "base": first_row.get("Left Base"),
                },
                {
                    "label": first_row.get("Right Label", "Group 2"),
                    "base": first_row.get("Right Base"),
                },
            ]
        )
    if not group_results:
        group_results = _normalize_topline_groups(
            [
                {"label": "Group 1", "base": None},
                {"label": "Group 2", "base": None},
            ]
        )
    pair_results = _normalize_topline_pairs(list(first_row.get("Comparison Pairs", []) or []), group_results)
    pair_count = len(pair_results)
    group_count = max(len(group_results), 1)
    _set_topline_columns(worksheet, group_count, pair_count)
    group_columns = [
        {
            "group": group,
            "column": 3 + position,
        }
        for position, group in enumerate(group_results)
    ]
    control_column = 3
    paired_columns = [
        {
            "pair": pair,
            "group_column": 4 + (position * 2),
            "lift_column": 5 + (position * 2),
        }
        for position, pair in enumerate(pair_results)
    ]
    note_base_column = (4 + (pair_count * 2)) if pair_count else 3 + group_count
    notes_start_column = note_base_column + 1
    note_column_count = max(pair_count, 1)
    end_column = notes_start_column + note_column_count - 1
    group_lookup = {group.get("index"): group for group in group_results}
    control_index = pair_results[0].get("left_index", group_results[0].get("index", 0)) if pair_results else group_results[0].get("index", 0)
    control_group = group_lookup.get(control_index, group_results[0])

    current_row = 1
    worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=end_column)
    title_cell = worksheet.cell(row=current_row, column=1, value="Viral Nation | Topline")
    _apply_header_style(title_cell, VN_BLACK)
    current_row = 2
    _write_version_stamp(worksheet, row=current_row, start_column=max(3, end_column - 2), end_column=end_column)
    worksheet.cell(row=current_row, column=2, value="Observations:")
    _apply_body_style(worksheet.cell(row=current_row, column=2), bold=True)
    for note in list(getattr(topline_sheet, "footnotes", []) or [])[:5]:
        current_row += 1
        worksheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=end_column)
        note_cell = worksheet.cell(row=current_row, column=2, value=note)
        _apply_body_style(note_cell, fill_color=VN_LIGHT_GRAY, wrap=True)
    current_row = 9

    results_cell = worksheet.cell(row=current_row, column=3, value="RESULTS")
    _apply_header_style(results_cell, VN_RED)
    current_row = 11
    comparison_variable_label = "Comparison"
    if rows:
        comparison_variable_label = rows[0].get("Comparison Variable", comparison_variable_label)

    header_values = {
        (11, 2): comparison_variable_label,
        (12, 2): "Base Size",
        (11, note_base_column): "Note Base",
    }
    if paired_columns:
        # Topline columns interleave each test group with its Control lift,
        # then split notes by Control-vs-group comparison.
        header_values[(11, control_column)] = control_group.get("label") or "Control"
        for position, paired_column in enumerate(paired_columns):
            pair = paired_column["pair"]
            right_label = pair.get("right_label", "Test")
            left_label = pair.get("left_label", "Control")
            header_values[(11, paired_column["group_column"])] = right_label
            header_values[(11, paired_column["lift_column"])] = f"{right_label} vs {left_label} Lift"
            header_values[(12, paired_column["lift_column"])] = "Lift"
            note_column = notes_start_column + position
            header_values[(11, note_column)] = f"{left_label} vs {right_label} Notes"
            header_values[(12, note_column)] = "Notes"
    else:
        for group_column in group_columns:
            group = group_column["group"]
            header_values[(11, group_column["column"])] = group.get("label") or "Group"
        header_values[(11, notes_start_column)] = "Notes"
        header_values[(12, notes_start_column)] = "Notes"

    for (row_index, column_index), value in header_values.items():
        cell = worksheet.cell(row=row_index, column=column_index, value=value)
        _apply_body_style(cell, bold=True, fill_color=VN_LIGHT_GRAY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if not rows:
        worksheet.merge_cells(start_row=14, start_column=2, end_row=14, end_column=end_column)
        empty_cell = worksheet.cell(
            row=14,
            column=2,
            value="No topline rows were generated for the current project setup.",
        )
        _apply_body_style(empty_cell, fill_color=VN_LIGHT_GRAY)
        worksheet.freeze_panes = "A4"
        return

    if paired_columns:
        base_cell = worksheet.cell(row=12, column=control_column, value=control_group.get("base"))
        _apply_body_style(base_cell)
        base_cell.alignment = Alignment(horizontal="center", vertical="top")
        for paired_column in paired_columns:
            pair = paired_column["pair"]
            right_group = group_lookup.get(pair.get("right_index"), {})
            base_cell = worksheet.cell(row=12, column=paired_column["group_column"], value=right_group.get("base"))
            _apply_body_style(base_cell)
            base_cell.alignment = Alignment(horizontal="center", vertical="top")
    else:
        for group_column in group_columns:
            group = group_column["group"]
            base_cell = worksheet.cell(row=12, column=group_column["column"], value=group.get("base"))
            _apply_body_style(base_cell)
            base_cell.alignment = Alignment(horizontal="center", vertical="top")
    note_base_header = worksheet.cell(row=12, column=note_base_column, value="")
    _apply_body_style(note_base_header, fill_color=VN_LIGHT_GRAY)

    current_row = 13
    for row in rows:
        row_label = row.get("Topline Label") or row.get("Variable") or row.get("Question", "")
        response = row.get("Response", "")
        if response:
            row_label = f"{row_label} | {response}"
        label_cell = worksheet.cell(row=current_row, column=2, value=row_label)
        _apply_body_style(label_cell, wrap=True)

        active_group_results = _normalize_topline_groups(list(row.get("Group Results", []) or []))
        if not active_group_results:
            active_group_results = _normalize_topline_groups(
                [
                    {"pct": row.get("Left %"), "sig": row.get("Left Sig", "")},
                    {"pct": row.get("Right %"), "sig": row.get("Right Sig", "")},
                ]
            )
        active_group_lookup = {group.get("index"): group for group in active_group_results}
        active_pair_results = _normalize_topline_pairs(
            list(row.get("Comparison Pairs", []) or []),
            active_group_results,
        )
        active_pair_lookup = {pair.get("note_key"): pair for pair in active_pair_results}

        if paired_columns:
            active_control_group = active_group_lookup.get(control_index)
            if active_control_group is None:
                active_control_group = active_group_results[0] if active_group_results else {}
            control_percentage = active_control_group.get("pct")
            control_sig_text = str(active_control_group.get("sig", "") or "")
            control_display = f"{control_percentage:.0%}{control_sig_text}" if control_percentage is not None else ""
            pct_cell = worksheet.cell(row=current_row, column=control_column, value=control_display)
            _apply_body_style(pct_cell)
            pct_cell.alignment = Alignment(horizontal="center", vertical="top")

            for position, paired_column in enumerate(paired_columns):
                header_pair = paired_column["pair"]
                pair = active_pair_lookup.get(header_pair.get("note_key"))
                if pair is None and position < len(active_pair_results):
                    pair = active_pair_results[position]
                pair = pair or header_pair
                right_group = active_group_lookup.get(pair.get("right_index")) or active_group_lookup.get(
                    header_pair.get("right_index")
                )
                if right_group is None and position + 1 < len(active_group_results):
                    right_group = active_group_results[position + 1]
                right_group = right_group or {}
                percentage = right_group.get("pct", pair.get("right_pct"))
                sig_text = str(right_group.get("sig", "") or "")
                display_value = f"{percentage:.0%}{sig_text}" if percentage is not None else ""
                pct_cell = worksheet.cell(row=current_row, column=paired_column["group_column"], value=display_value)
                _apply_body_style(pct_cell)
                pct_cell.alignment = Alignment(horizontal="center", vertical="top")

                lift_value = pair.get("lift")
                lift_cell = worksheet.cell(row=current_row, column=paired_column["lift_column"], value=lift_value)
                _apply_body_style(lift_cell, fill_color=VN_LIGHT_GRAY)
                if lift_value is not None:
                    lift_cell.number_format = "0%"
                _apply_topline_delta_style(lift_cell, str(pair.get("sig_direction", "") or ""))
        else:
            for group_column in group_columns:
                group_index = group_column["group"].get("index")
                active_group = active_group_lookup.get(group_index, {})
                percentage = active_group.get("pct")
                sig_text = str(active_group.get("sig", "") or "")
                display_value = f"{percentage:.0%}{sig_text}" if percentage is not None else ""
                pct_cell = worksheet.cell(row=current_row, column=group_column["column"], value=display_value)
                _apply_body_style(pct_cell)
                pct_cell.alignment = Alignment(horizontal="center", vertical="top")

        note_base_cell = worksheet.cell(
            row=current_row,
            column=note_base_column,
            value=row.get("Note Base") or "Total Answering",
        )
        _apply_body_style(note_base_cell, fill_color=VN_LIGHT_GRAY)
        note_base_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        pair_notes = row.get("Pair Notes", {}) or {}
        if paired_columns:
            for position, paired_column in enumerate(paired_columns):
                pair = paired_column["pair"]
                note_value = pair_notes.get(pair.get("note_key"), "")
                if not note_value and pair_count == 1:
                    note_value = row.get("Notes", "")
                notes_cell = worksheet.cell(row=current_row, column=notes_start_column + position, value=note_value)
                _apply_body_style(notes_cell, wrap=True)
        else:
            notes_cell = worksheet.cell(row=current_row, column=notes_start_column, value=row.get("Notes", ""))
            _apply_body_style(notes_cell, wrap=True)
        current_row += 1

    worksheet.freeze_panes = "B11"


def _get_metric_flag(source, fallback, name: str, default: bool) -> bool:
    """Resolve one table metric flag with sheet-level fallback."""
    if hasattr(source, name):
        return bool(getattr(source, name))
    if hasattr(fallback, name):
        return bool(getattr(fallback, name))
    return default


def _build_metric_row_kinds(
    include_percentage: bool,
    include_stat_testing: bool,
    include_n_count: bool,
    notation_location: str,
) -> list[str]:
    """Return the per-response metric rows that should be written."""
    row_kinds: list[str] = []
    if include_percentage:
        row_kinds.append("percentage")
    if include_stat_testing and (notation_location == "below_metric" or not include_percentage):
        row_kinds.append("sig")
    if include_n_count:
        row_kinds.append("n")
    return row_kinds


def _format_section_metric_title(
    section_label: str,
    base_label: str,
    include_percentage: bool,
    include_stat_testing: bool,
    include_n_count: bool,
    lift_enabled: bool,
) -> str:
    """Build the section header based on the selected output metrics."""
    metric_labels: list[str] = []
    if include_percentage:
        metric_labels.append("%")
    if lift_enabled:
        metric_labels.append("Lift")
    if include_stat_testing:
        metric_labels.append("Stat testing")
    if include_n_count:
        metric_labels.append("N Count")
    metric_text = " / ".join(metric_labels) if metric_labels else "No metrics"
    return f"{section_label} {metric_text} (Base: {base_label})"


def _write_banner_sheet(
    workbook,
    sheet,
    include_lift: bool = False,
) -> None:
    """Write one banner worksheet in an analyst-friendly table format."""
    table_groups = [table.groups for table in sheet.tables if getattr(table, "groups", None)]
    max_group_count = max([len(sheet.groups), *[len(groups) for groups in table_groups], 1])
    max_lift_count = 0
    if include_lift:
        lift_sources = []
        if sheet.groups:
            lift_sources.append((sheet, sheet.groups))
        lift_sources.extend((table, table.groups) for table in sheet.tables if table.groups)
        for source, groups in lift_sources:
            lift_pairs, _ = _build_banner_lift_pairs(source, list(groups))
            max_lift_count = max(max_lift_count, len(lift_pairs))

    worksheet = workbook.create_sheet(title=str(sheet.name)[:31] or "Sheet1")
    _set_sheet_columns(worksheet, max_group_count, max_lift_count)

    current_row = 1
    max_end_column = 3 + (max_group_count * 2) + (max_lift_count * 2)
    worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_end_column)
    title_cell = worksheet.cell(row=current_row, column=1, value=sheet.banner_name)
    _apply_header_style(title_cell, VN_BLACK)
    current_row += 1
    _write_version_stamp(
        worksheet,
        row=current_row,
        start_column=max(3, max_end_column - 2),
        end_column=max_end_column,
    )
    worksheet.cell(row=current_row, column=1, value="filtered by")
    _apply_body_style(worksheet.cell(row=current_row, column=1))
    current_row += 1
    worksheet.cell(row=current_row, column=1, value="FILTER: ALL")
    _apply_body_style(worksheet.cell(row=current_row, column=1))
    current_row += 2

    def _write_one_table(
        current_row: int,
        table,
        active_banner_name: str,
        active_levels: list[str],
        active_groups: list[dict[str, object]],
        footnotes: list[str],
        active_level_labels: dict[str, str] | None = None,
    ) -> int:
        lift_source = type(
            "obj",
            (),
            {
                "levels": active_levels,
                "comparison_pairs": list(getattr(table, "comparison_pairs", []) or getattr(sheet, "comparison_pairs", []) or []),
            },
        )
        lift_pairs, lift_footnote = _build_banner_lift_pairs(lift_source, active_groups) if include_lift else ([], "")
        table_include_percentage = _get_metric_flag(table, sheet, "include_percentage", True)
        table_include_n_count = _get_metric_flag(table, sheet, "include_n_count", False)
        table_include_stat_testing = _get_metric_flag(table, sheet, "include_stat_testing", True)
        table_notation_location = normalize_text(
            getattr(table, "notation_location", getattr(sheet, "notation_location", "appended_to_metric"))
        ) or "appended_to_metric"
        lift_enabled = bool(include_lift and lift_pairs)
        metric_row_kinds = _build_metric_row_kinds(
            table_include_percentage,
            table_include_stat_testing,
            table_include_n_count,
            table_notation_location,
        )
        section_group_count = max(len(active_groups), 1)
        left_data_start_column = 3
        left_lift_start_column = left_data_start_column + section_group_count
        separator_column = left_lift_start_column + (len(lift_pairs) if lift_enabled else 0)
        right_data_start_column = separator_column + 1
        right_lift_start_column = right_data_start_column + section_group_count
        left_data_columns = [left_data_start_column + index for index in range(len(active_groups))]
        right_data_columns = [right_data_start_column + index for index in range(len(active_groups))]
        left_lift_columns = [left_lift_start_column + index for index in range(len(lift_pairs))] if lift_enabled else []
        right_lift_columns = [right_lift_start_column + index for index in range(len(lift_pairs))] if lift_enabled else []
        table_end_column = (
            right_lift_columns[-1] if right_lift_columns else
            (right_data_columns[-1] if right_data_columns else 3)
        )

        active_level_labels = active_level_labels or {}
        banner_descriptor = " > ".join(
            _format_level_label(level, active_level_labels) for level in active_levels
        ) if active_levels else active_banner_name or "All Tables"
        if active_groups:
            worksheet.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=table_end_column)
            descriptor_cell = worksheet.cell(row=current_row, column=3, value=banner_descriptor)
            _apply_header_style(descriptor_cell, VN_ORANGE)
        current_row += 1

        level_rows = max(1, len(active_levels))
        total_group_indexes = [index for index, group in enumerate(active_groups) if group["label"] == "Total"]
        non_total_indexes = [index for index, group in enumerate(active_groups) if group["label"] != "Total"]

        section_header_row = current_row
        if active_groups:
            worksheet.merge_cells(
                start_row=section_header_row,
                start_column=left_data_start_column,
                end_row=section_header_row,
                end_column=(left_lift_columns[-1] if left_lift_columns else left_data_columns[-1]),
            )
            left_title_cell = worksheet.cell(
                row=section_header_row,
                column=left_data_start_column,
                value=_format_section_metric_title(
                    "Total Sample",
                    "Total Sample",
                    table_include_percentage,
                    table_include_stat_testing,
                    table_include_n_count,
                    lift_enabled,
                ),
            )
            _apply_header_style(left_title_cell, VN_YELLOW, font_color=VN_BLACK)
            worksheet.merge_cells(
                start_row=section_header_row,
                start_column=right_data_start_column,
                end_row=section_header_row,
                end_column=(right_lift_columns[-1] if right_lift_columns else right_data_columns[-1]),
            )
            right_title_cell = worksheet.cell(
                row=section_header_row,
                column=right_data_start_column,
                value=_format_section_metric_title(
                    "Total Answering",
                    "Total Answering",
                    table_include_percentage,
                    table_include_stat_testing,
                    table_include_n_count,
                    lift_enabled,
                ),
            )
            _apply_header_style(right_title_cell, VN_YELLOW, font_color=VN_BLACK)
        current_row += 1

        def _write_group_header_block(data_columns: list[int]) -> None:
            if total_group_indexes:
                total_column = data_columns[total_group_indexes[0]]
                worksheet.merge_cells(
                    start_row=current_row,
                    start_column=total_column,
                    end_row=current_row + level_rows - 1,
                    end_column=total_column,
                )
                total_cell = worksheet.cell(row=current_row, column=total_column, value="Total")
                _apply_body_style(total_cell, bold=True, fill_color=VN_LIGHT_GRAY)
                total_cell.alignment = Alignment(horizontal="center", vertical="center")
            if active_levels and non_total_indexes:
                for level_index, level_name in enumerate(active_levels):
                    header_row = current_row + level_index
                    start_group = None
                    previous_value = None
                    for position, group_index in enumerate(non_total_indexes):
                        group = active_groups[group_index]
                        current_value = normalize_text(group.get("display_values", {}).get(level_name))
                        if not current_value:
                            current_value = normalize_text(group.get("values", {}).get(level_name))
                        if start_group is None:
                            start_group = group_index
                            previous_value = current_value
                            continue
                        if current_value != previous_value:
                            start_column = data_columns[start_group]
                            end_column = data_columns[non_total_indexes[position - 1]]
                            worksheet.merge_cells(
                                start_row=header_row,
                                start_column=start_column,
                                end_row=header_row,
                                end_column=end_column,
                            )
                            merged_cell = worksheet.cell(row=header_row, column=start_column, value=previous_value)
                            _apply_body_style(merged_cell, bold=True)
                            merged_cell.alignment = Alignment(horizontal="center", vertical="center")
                            start_group = group_index
                            previous_value = current_value
                    if start_group is not None:
                        start_column = data_columns[start_group]
                        end_column = data_columns[non_total_indexes[-1]]
                        worksheet.merge_cells(
                            start_row=header_row,
                            start_column=start_column,
                            end_row=header_row,
                            end_column=end_column,
                        )
                        merged_cell = worksheet.cell(row=header_row, column=start_column, value=previous_value)
                        _apply_body_style(merged_cell, bold=True)
                        merged_cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                header_row = current_row
                for column_index, group in zip(data_columns, active_groups):
                    group_cell = worksheet.cell(row=header_row, column=column_index, value=group["label"])
                    _apply_body_style(group_cell, bold=(group["label"] == "Total"), fill_color=VN_LIGHT_GRAY if group["label"] == "Total" else None)
                    group_cell.alignment = Alignment(horizontal="center", vertical="center")

        _write_group_header_block(left_data_columns)
        _write_group_header_block(right_data_columns)

        header_end_row = current_row + level_rows - 1
        sig_row = header_end_row
        if table_include_stat_testing:
            sig_row = header_end_row + 1
            for data_columns in [left_data_columns, right_data_columns]:
                for index, group in enumerate(active_groups):
                    column_index = data_columns[index]
                    if group["label"] != "Total":
                        sig_letter = chr(64 + index) if index < 27 else ""
                        sig_cell = worksheet.cell(row=sig_row, column=column_index, value=sig_letter)
                        _apply_body_style(sig_cell)
                        sig_cell.alignment = Alignment(horizontal="center", vertical="center")
        if lift_enabled:
            for column_index, pair in zip(left_lift_columns, lift_pairs):
                worksheet.merge_cells(start_row=current_row, start_column=column_index, end_row=sig_row, end_column=column_index)
                lift_header_cell = worksheet.cell(row=current_row, column=column_index, value=f"{pair['parent_label']} Lift")
                _apply_body_style(lift_header_cell, bold=True, fill_color=VN_LIGHT_GRAY, wrap=True)
                lift_header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for column_index, pair in zip(right_lift_columns, lift_pairs):
                worksheet.merge_cells(start_row=current_row, start_column=column_index, end_row=sig_row, end_column=column_index)
                lift_header_cell = worksheet.cell(row=current_row, column=column_index, value=f"{pair['parent_label']} Lift")
                _apply_body_style(lift_header_cell, bold=True, fill_color=VN_LIGHT_GRAY, wrap=True)
                lift_header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        current_row = sig_row + 3

        total_base_section = next((section for section in table.sections if section["label"] == "Total Base"), None)
        answering_section = next((section for section in table.sections if section["label"] == "Total Answering"), None)
        if not total_base_section or not answering_section:
            return current_row

        response_block_height = len(total_base_section["rows"]) * len(metric_row_kinds)
        question_label_row = current_row + 2
        question_end_row = question_label_row + max(response_block_height - 1, 0)
        if response_block_height > 1:
            worksheet.merge_cells(start_row=question_label_row, start_column=1, end_row=question_end_row, end_column=1)
        question_cell = worksheet.cell(row=question_label_row, column=1, value=table.question_label)
        _apply_header_style(question_cell, VN_RED)
        question_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        count_label_cell = worksheet.cell(row=current_row, column=2, value="Base Sizes")
        _apply_body_style(count_label_cell, bold=True, fill_color=VN_LIGHT_GRAY)
        for column_index, denominator in zip(left_data_columns, total_base_section["base_denominators"]):
            cell = worksheet.cell(row=current_row, column=column_index, value=denominator)
            _apply_body_style(cell, bold=True, fill_color=VN_LIGHT_GRAY)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_index, denominator in zip(right_data_columns, answering_section["base_denominators"]):
            cell = worksheet.cell(row=current_row, column=column_index, value=denominator)
            _apply_body_style(cell, bold=True, fill_color=VN_LIGHT_GRAY)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_index, pair in zip(left_lift_columns, lift_pairs):
            lift_base_value = (
                total_base_section["base_denominators"][pair["left_index"]]
                + total_base_section["base_denominators"][pair["right_index"]]
            )
            cell = worksheet.cell(row=current_row, column=column_index, value=lift_base_value)
            _apply_body_style(cell, bold=True, fill_color=VN_LIGHT_GRAY)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_index, pair in zip(right_lift_columns, lift_pairs):
            lift_base_value = (
                answering_section["base_denominators"][pair["left_index"]]
                + answering_section["base_denominators"][pair["right_index"]]
            )
            cell = worksheet.cell(row=current_row, column=column_index, value=lift_base_value)
            _apply_body_style(cell, bold=True, fill_color=VN_LIGHT_GRAY)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 2

        for total_row, answering_row in zip(total_base_section["rows"], answering_section["rows"]):
            label_fill = VN_LIGHT_GRAY if total_row.get("kind") == "net" else None
            for metric_index, row_kind in enumerate(metric_row_kinds):
                response_cell = worksheet.cell(
                    row=current_row,
                    column=2,
                    value=total_row["label"] if metric_index == 0 else "",
                )
                _apply_body_style(
                    response_cell,
                    bold=bool(total_row.get("kind") == "net") and metric_index == 0,
                    fill_color=label_fill if metric_index else None,
                    wrap=True,
                )
                response_cell.alignment = Alignment(vertical="center", wrap_text=True)

                for data_columns, section_row in [(left_data_columns, total_row), (right_data_columns, answering_row)]:
                    if row_kind == "percentage":
                        for column_index, percentage, sig_text in zip(data_columns, section_row["percentages"], section_row["sig_letters"]):
                            display_value = ""
                            if percentage is not None:
                                sig_suffix = normalize_text(sig_text) if table_include_stat_testing and table_notation_location != "below_metric" else ""
                                display_value = f"{percentage:.0%}{sig_suffix}"
                            metric_cell = worksheet.cell(row=current_row, column=column_index, value=display_value)
                            _apply_body_style(metric_cell, fill_color=label_fill)
                            metric_cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif row_kind == "sig":
                        for column_index, sig_text in zip(data_columns, section_row["sig_letters"]):
                            metric_cell = worksheet.cell(row=current_row, column=column_index, value=normalize_text(sig_text))
                            _apply_body_style(metric_cell, fill_color=label_fill)
                            metric_cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif row_kind == "n":
                        for column_index, count in zip(data_columns, section_row["counts"]):
                            metric_cell = worksheet.cell(row=current_row, column=column_index, value=count)
                            _apply_body_style(metric_cell, fill_color=label_fill)
                            metric_cell.alignment = Alignment(horizontal="center", vertical="center")

                if lift_enabled:
                    write_lift_values = row_kind == "percentage" or (not table_include_percentage and metric_index == 0)
                    for column_index, pair in zip(left_lift_columns, lift_pairs):
                        lift_value = (
                            _format_lift_display(total_row["percentages"][pair["left_index"]], total_row["percentages"][pair["right_index"]])
                            if write_lift_values
                            else ""
                        )
                        lift_cell = worksheet.cell(row=current_row, column=column_index, value=lift_value)
                        _apply_body_style(lift_cell, fill_color=label_fill)
                        lift_cell.alignment = Alignment(horizontal="center", vertical="center")
                    for column_index, pair in zip(right_lift_columns, lift_pairs):
                        lift_value = (
                            _format_lift_display(answering_row["percentages"][pair["left_index"]], answering_row["percentages"][pair["right_index"]])
                            if write_lift_values
                            else ""
                        )
                        lift_cell = worksheet.cell(row=current_row, column=column_index, value=lift_value)
                        _apply_body_style(lift_cell, fill_color=label_fill)
                        lift_cell.alignment = Alignment(horizontal="center", vertical="center")

                current_row += 1

        if include_lift and not lift_enabled and lift_footnote:
            footnotes = [*footnotes, lift_footnote]
        if footnotes:
            for note in footnotes:
                worksheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=table_end_column)
                footnote_cell = worksheet.cell(row=current_row, column=2, value=note)
                _apply_body_style(footnote_cell, fill_color=VN_LIGHT_GRAY, wrap=True)
                footnote_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                current_row += 1
        _paint_separator_column(worksheet, separator_column, section_header_row, current_row - 1)
        return current_row + 2

    previous_banner_name = None
    if sheet.groups:
        for table in sheet.tables:
            current_row = _write_one_table(
                current_row,
                table,
                sheet.banner_name,
                sheet.levels,
                list(sheet.groups),
                list(sheet.footnotes),
                dict(getattr(sheet, "level_labels", {}) or {}),
            )
    else:
        for table in sheet.tables:
            active_banner_name = getattr(table, "banner_name", sheet.banner_name)
            active_groups = list(getattr(table, "groups", []))
            table_end_column = max_end_column
            if active_banner_name != previous_banner_name:
                worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=table_end_column)
                banner_section_cell = worksheet.cell(row=current_row, column=1, value=f"Banner: {active_banner_name}")
                _apply_header_style(banner_section_cell, VN_BLACK)
                current_row += 2
                previous_banner_name = active_banner_name
            current_row = _write_one_table(
                current_row,
                table,
                active_banner_name,
                list(getattr(table, "levels", [])),
                active_groups,
                list(getattr(table, "footnotes", [])),
                dict(getattr(table, "level_labels", {}) or {}),
            )

    worksheet.freeze_panes = "D7"


def export_workbook_to_excel_bytes(
    workbook_package: dict,
    uploaded_filename: str | None = None,
) -> bytes:
    """Convert the generated workbook package into branded Excel bytes.

    Inputs:
        workbook_package: Structured workbook content built by the table
        generator.
        uploaded_filename: Optional source filename, used only by the calling
        layer for naming the final download.

    Outputs:
        In-memory Excel bytes ready for a Streamlit download button.
    """
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    topline_sheet = workbook_package.get("topline_sheet")
    if topline_sheet is not None:
        _write_topline_sheet(workbook, topline_sheet)

    for sheet in workbook_package.get("sheets", []):
        _write_banner_sheet(
            workbook,
            sheet,
            include_lift=bool(workbook_package.get("include_lift", False)),
        )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_tables_to_excel_bytes(tables: dict) -> bytes:
    """Compatibility export helper for older placeholder dataframe payloads.

    Inputs:
        A mapping of sheet names to pandas dataframes.

    Outputs:
        In-memory Excel bytes for legacy placeholder flows.
    """
    from pandas import ExcelWriter

    output = BytesIO()
    with ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, dataframe in tables.items():
            safe_name = str(sheet_name)[:31] or "Sheet1"
            dataframe.to_excel(writer, sheet_name=safe_name, index=False)
    return output.getvalue()
